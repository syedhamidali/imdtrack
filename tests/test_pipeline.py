"""Tests for the publish pipeline: storage round-trip, validation, and the
fail-safe "don't clobber good data with a broken parse" guarantee.
"""

from datetime import datetime

import openpyxl
import pandas as pd
import pytest

from imdtrack import store, validate
from imdtrack.parse import parse_workbook
from imdtrack.pipeline import build

HEADER_14 = [
    "Serial Number of system during year",
    "Basin of origin",
    "Name",
    "Date(DD-MM-YYYY)",
    "Time (UTC)",
    "Latitude (lat.)",
    "Longitude (lon.)",
    'CI No [or "T. No"]',
    "Estimated Central Pressure (hPa)",
    "Maximum Sustained Surface Wind (kt) ",
    "Pressure Drop (hPa)",
    "Grade (text)",
    "Outermost closed isobar (hPa)",
    "Diameter/Size of outermost closed isobar",
]


def _storm_rows(serial, name, day):
    d = datetime(2020, 5, day)
    return [
        [serial, "BOB", name, d, "0000", 10.4, 87.0, 1.5, "1000", "25", "3", "D", "", ""],
        [None, "BOB", name, d, "0300", 10.7, 86.5, 2.0, "998", "30", "5", "DD", "", ""],
        [None, "BOB", name, d, "0600", 11.0, 86.0, 2.5, "996", "35", "7", "CS", "", ""],
    ]


def _write_workbook(tmp_path, n_storms=3, name="wb.xlsx"):
    rows = [HEADER_14]
    for i in range(1, n_storms + 1):
        rows.extend(_storm_rows(i, f"STORM{i}", 10 + i))
        rows.append([])  # separator
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="2020")
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# store round-trip
# --------------------------------------------------------------------------- #


def test_store_roundtrip_and_manifest(tmp_path):
    wb = _write_workbook(tmp_path)
    frames = parse_workbook(wb)
    data_dir = tmp_path / "data"

    manifest = store.write_dataset(frames, data_dir, source_url="http://x", source_sha256="abc123")

    assert store.dataset_exists(data_dir)
    assert manifest.n_obs == len(frames["observations"])
    assert manifest.n_storms == 3
    assert manifest.year_min == 2020 and manifest.year_max == 2020

    back = store.read_dataset(data_dir)
    assert len(back["observations"]) == len(frames["observations"])
    # ordered categorical is restored
    assert back["observations"]["grade"].cat.ordered

    read_m = store.read_manifest(data_dir)
    assert read_m.source_sha256 == "abc123"


# --------------------------------------------------------------------------- #
# validation
# --------------------------------------------------------------------------- #


def test_validate_rejects_empty():
    empty = {
        "observations": pd.DataFrame(),
        "remarks": pd.DataFrame(),
        "storms": pd.DataFrame(),
    }
    with pytest.raises(validate.ValidationError):
        validate.validate_frames(empty)


def test_validate_rejects_out_of_range_positions(tmp_path):
    wb = _write_workbook(tmp_path)
    frames = parse_workbook(wb)
    frames["observations"]["lat"] = 999.0  # corrupt every latitude
    with pytest.raises(validate.ValidationError):
        validate.validate_frames(frames)


def test_validate_against_previous_rejects_shrink(tmp_path):
    wb = _write_workbook(tmp_path, n_storms=3)
    frames = parse_workbook(wb)
    prev = store.Manifest(
        format=1,
        source_url="x",
        source_sha256="old",
        generated_at="",
        n_storms=100,
        n_obs=10000,
        n_remarks=0,
        year_min=1982,
        year_max=2020,
    )
    with pytest.raises(validate.ValidationError):
        validate.validate_against_previous(frames, prev)


# --------------------------------------------------------------------------- #
# pipeline build (patch fetch so no network is used)
# --------------------------------------------------------------------------- #


def _patch_fetch(monkeypatch, wb_path, sha):
    from imdtrack import fetch as _fetch
    from imdtrack import pipeline

    def fake_fetch_workbook(url, cache_dir=None, force=False, timeout=120.0):
        return _fetch.FetchResult(path=wb_path, changed=True, sha256=sha, from_cache=False)

    monkeypatch.setattr(pipeline._fetch, "fetch_workbook", fake_fetch_workbook)


def test_build_initial_then_unchanged(tmp_path, monkeypatch):
    wb = _write_workbook(tmp_path)
    data_dir = tmp_path / "data"

    _patch_fetch(monkeypatch, wb, sha="sha-v1")
    r1 = build(data_dir, url="http://x")
    assert r1.changed and r1.reason == "initial build"
    assert store.dataset_exists(data_dir)

    # Same sha -> no change, no re-parse.
    r2 = build(data_dir, url="http://x")
    assert not r2.changed and r2.reason == "source unchanged"


def test_build_refuses_to_clobber_on_bad_parse(tmp_path, monkeypatch):
    # 1) Publish a good dataset.
    good_wb = _write_workbook(tmp_path, n_storms=3, name="good.xlsx")
    data_dir = tmp_path / "data"
    _patch_fetch(monkeypatch, good_wb, sha="sha-good")
    build(data_dir, url="http://x")
    good_manifest = store.read_manifest(data_dir)
    assert good_manifest.n_storms == 3

    # 2) A "new" IMD file that parses to nothing usable (no header -> empty).
    bad_wb = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="2021")
    ws.append(["garbage", "with", "no", "recognisable", "header"])
    wb.save(bad_wb)

    _patch_fetch(monkeypatch, bad_wb, sha="sha-bad")
    with pytest.raises(validate.ValidationError):
        build(data_dir, url="http://x")

    # 3) The committed dataset must be exactly the good one, untouched.
    after = store.read_manifest(data_dir)
    assert after.source_sha256 == "sha-good"
    assert after.n_storms == 3
    assert store.read_dataset(data_dir)["storms"].shape[0] == 3
