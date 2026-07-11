"""Parse the IMD Best Track workbook into tidy pandas structures.

The workbook has one sheet per year.  Each sheet contains several storms, one
after another; a storm's ``serial`` number appears only on its first row, and
free-text landfall / weakening notes are interleaved between the 3-hourly track
rows.  :func:`parse_workbook` untangles all of that into two frames:

* ``observations`` - one tidy row per 3-hourly fix (the "track").
* ``remarks``      - the free-text notes, linked back to their storm.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Union

import pandas as pd

from . import _schema as S

PathLike = Union[str, Path]


def _iter_year_sheets(wb):
    for name in wb.sheetnames:
        name_s = str(name).strip()
        if name_s.isdigit() and 1800 < int(name_s) < 2200:
            yield int(name_s), wb[name]


def _remark_text(row) -> str:
    """Join the long free-text fragments of a non-track row into one string."""
    parts = []
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip()
        if len(s) > 12 and any(ch.isalpha() for ch in s):
            parts.append(s)
    return " ".join(parts)


def parse_sheet(year: int, ws) -> tuple[list[dict], list[dict]]:
    rows = ws.iter_rows(values_only=True)

    # Locate the header row (it may be preceded by note rows).
    colmap = None
    for row in rows:
        if S.is_header_row(row):
            colmap = S.detect_columns(row)
            break
    if colmap is None:
        return [], []

    def get(row, field):
        idx = colmap.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    observations: list[dict] = []
    remarks: list[dict] = []

    cur_serial = None
    cur_basin = None
    cur_name = None
    cur_date = None
    storm_no = 0
    separator_seen = True  # start of sheet counts as a storm boundary

    for row in rows:
        if not any(c is not None for c in row):
            separator_seen = True  # blank rows separate storms
            continue

        serial = get(row, "serial")
        serial_val = None
        if serial is not None:
            try:
                serial_val = int(float(serial))
            except (TypeError, ValueError):
                serial_val = None

        # A serial appears only on a storm's first row (merged cell), so its
        # presence after a separator — or with a changed value — marks a new
        # storm. IMD restarts the serial per basin in older years, so keying on
        # (year, serial) merges e.g. ARB #1 with BOB #1; track a running storm
        # number per sheet instead so every storm gets a unique id.
        if serial_val is not None and (separator_seen or serial_val != cur_serial):
            storm_no += 1
            cur_serial = serial_val
            cur_name = None  # don't let the previous storm's fields forward-fill in
            cur_basin = None
            cur_date = None
        elif serial_val is not None:
            cur_serial = serial_val
        separator_seen = False

        basin = S.normalize_basin(get(row, "basin"))
        if basin is not None:
            cur_basin = basin
        name = S.normalize_name(get(row, "name"))
        if name is not None:
            cur_name = name

        # The Date column is a *merged* cell in the workbook: it holds the date
        # only on the first (usually 00:00) row of each day and is None on the
        # 3-hourly rows below it. It is also stored inconsistently (real datetime
        # in most sheets, day-first text like "16-06-2008" in others). Parse both
        # forms and forward-fill — like serial/name/basin — so no fix is dropped.
        date_cell = S.parse_date(get(row, "date"))
        if date_cell is not None:
            cur_date = date_cell

        minutes = S.parse_time(get(row, "time"))
        lat, lon = S.parse_latlon(get(row, "lat"), get(row, "lon"))

        is_track = (
            storm_no > 0
            and isinstance(cur_date, datetime)
            and minutes is not None
            and lat == lat  # not NaN
            and lon == lon
        )

        if is_track:
            observations.append(
                {
                    "year": year,
                    "storm_no": storm_no,
                    "serial": cur_serial,
                    "basin": cur_basin,
                    "name": cur_name,
                    "time": cur_date + timedelta(minutes=minutes),
                    "lat": lat,
                    "lon": lon,
                    "ci_no": S.to_float(get(row, "ci_no")),
                    "pressure": S.to_float(get(row, "pressure")),
                    "wind": S.to_float(get(row, "wind")),
                    "pressure_drop": S.to_float(get(row, "pressure_drop")),
                    "grade": S.normalize_grade(get(row, "grade")),
                    "oci": S.to_float(get(row, "oci")),
                    "oci_diameter": S.to_float(get(row, "oci_diameter")),
                }
            )
        elif storm_no > 0:
            text = _remark_text(row)
            if text:
                remarks.append(
                    {
                        "year": year,
                        "storm_no": storm_no,
                        "serial": cur_serial,
                        "date": cur_date if isinstance(cur_date, datetime) else pd.NaT,
                        "remark": text,
                    }
                )

    return observations, remarks


def count_positional_rows(path: PathLike) -> dict[int, int]:
    """Count, per year, rows that carry a real fix — a numeric lat, numeric lon,
    and a valid HHMM time — under a storm.

    This is deliberately **independent of the date handling** (which is where the
    fiddly parsing lives): it answers "how many track fixes does the workbook
    contain?" so the pipeline can verify the parser captured all of them and
    never silently drops rows again.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        counts: dict[int, int] = {}
        for year, ws in _iter_year_sheets(wb):
            rows = ws.iter_rows(values_only=True)
            colmap = None
            for row in rows:
                if S.is_header_row(row):
                    colmap = S.detect_columns(row)
                    break
            if colmap is None:
                continue

            def get(row, field, colmap=colmap):
                idx = colmap.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            cur_serial = None
            n = 0
            for row in rows:
                if not any(c is not None for c in row):
                    continue
                serial = get(row, "serial")
                if serial is not None:
                    try:
                        cur_serial = int(float(serial))
                    except (TypeError, ValueError):
                        pass
                lat = S.to_float(get(row, "lat"))
                lon = S.to_float(get(row, "lon"))
                minutes = S.parse_time(get(row, "time"))
                if cur_serial is not None and lat == lat and lon == lon and minutes is not None:
                    n += 1
            counts[year] = n
    finally:
        wb.close()
    return counts


def _storm_id(year: int, storm_no: int) -> str:
    return f"{year}-{storm_no:03d}"


def parse_workbook(path: PathLike) -> dict[str, pd.DataFrame]:
    """Parse a workbook file into ``{"observations", "remarks", "storms"}`` frames."""
    import openpyxl  # deferred: only the pipeline parses raw workbooks.

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        all_obs: list[dict] = []
        all_rem: list[dict] = []
        for year, ws in _iter_year_sheets(wb):
            obs, rem = parse_sheet(year, ws)
            all_obs.extend(obs)
            all_rem.extend(rem)
    finally:
        wb.close()

    obs = pd.DataFrame(
        all_obs,
        columns=["year", "storm_no", "serial", "basin", "name", "time"]
        + [f for f in S.FIELDS if f in S.NUMERIC_FIELDS]
        + ["grade"],
    )
    if not obs.empty:
        obs.insert(0, "storm_id", [_storm_id(y, n) for y, n in zip(obs["year"], obs["storm_no"])])
        obs = obs.drop(columns="storm_no")
        obs["grade"] = pd.Categorical(obs["grade"], categories=S.GRADE_ORDER, ordered=True)
        obs["basin"] = obs["basin"].astype("category")
        obs = obs.sort_values(["storm_id", "time"], kind="stable").reset_index(drop=True)
        # observation index within each storm (0-based)
        obs["step"] = obs.groupby("storm_id", sort=False).cumcount()
        # Non-destructive QC: flag isolated position spikes and day/month-swapped
        # dates (both are source data-entry errors; values are left untouched).
        from . import qc

        obs["pos_suspect"] = qc.flag_positions(obs)
        obs["date_suspect"], _ = qc.flag_dates(obs)

    rem = pd.DataFrame(all_rem, columns=["year", "storm_no", "serial", "date", "remark"])
    if not rem.empty:
        rem.insert(0, "storm_id", [_storm_id(y, n) for y, n in zip(rem["year"], rem["storm_no"])])
        rem = rem.drop(columns="storm_no")

    storms = _summarize_storms(obs)
    return {"observations": obs, "remarks": rem, "storms": storms}


def _summarize_storms(obs: pd.DataFrame) -> pd.DataFrame:
    if obs.empty:
        return pd.DataFrame(
            columns=[
                "storm_id",
                "year",
                "serial",
                "name",
                "basin",
                "start_time",
                "end_time",
                "n_obs",
                "peak_grade",
                "max_wind",
                "min_pressure",
            ]
        )

    def peak_grade(s):
        s = s.dropna()
        return s.max() if len(s) else None

    g = obs.groupby("storm_id", sort=False)
    storms = g.agg(
        year=("year", "first"),
        serial=("serial", "first"),
        name=("name", "first"),
        basin=("basin", "first"),
        start_time=("time", "min"),
        end_time=("time", "max"),
        n_obs=("time", "size"),
        max_wind=("wind", "max"),
        min_pressure=("pressure", "min"),
    ).reset_index()
    storms["peak_grade"] = g["grade"].apply(peak_grade).values
    return storms
